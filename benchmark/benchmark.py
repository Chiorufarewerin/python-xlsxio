import sys
import timeit
import platform
from functools import partial


FILE_XLSX_BIG = "benchmark_big.xlsx"
FILE_XLSX_SMALL = "benchmark_small.xlsx"

with open(FILE_XLSX_BIG, 'rb') as f:
    FILE_XLSX_BIG_MEMORY = f.read()

with open(FILE_XLSX_SMALL, 'rb') as f:
    FILE_XLSX_SMALL_MEMORY = f.read()


READ_XLSXIO_SETUP = '''
import xlsxio
import datetime
types = (int, str, str, datetime.datetime, int, float, datetime.datetime, bool)
filename = {filename}
'''

READ_XLSXIO_STRING = '''
with xlsxio.XlsxioReader(filename) as reader:
    with reader.get_sheet(types=types if {with_types} else None, default_type={default_type}) as sheet:
        data = sheet.read_data()
'''


READ_OPENPYXL_SETUP = '''
import io
import openpyxl
filename = {filename}
'''

READ_OPENPYXL_STRING = '''
if isinstance(filename, bytes):
    filename = io.BytesIO(filename)
wb = openpyxl.load_workbook(filename=filename, read_only=True)
ws = wb.active
data = list(ws.iter_rows())
wb.close()
'''

READ_XLRD_SETUP = '''
import xlrd
filename = {filename}
if isinstance(filename, bytes):
    kwargs = {{'file_contents': filename}}
else:
    kwargs = {{'filename': filename}}
'''

READ_XLRD_STRING = '''
wb = xlrd.open_workbook(**kwargs)
ws = wb.sheet_by_index(0)
data = list(ws.get_rows())
'''

READ_SXL_SETUP = '''
import io
import sxl
filename = {filename}
'''

READ_SXL_STRING = '''
if isinstance(filename, bytes):
    filename = io.BytesIO(filename)
wb = sxl.Workbook(filename)
ws = wb.sheets[1]
data = list(ws.rows)
'''


READ_XLSX2CSV_SETUP = '''
import io
import csv
import xlsx2csv
filename = {filename}
'''

READ_XLSX2CSV_STRING = '''
if isinstance(filename, bytes):
    filename = io.BytesIO(filename)
data = io.StringIO()
xlsx2csv.Xlsx2csv(filename).convert(data)
data.seek(0)
data = list(csv.reader(data))
'''


def benchmark_one(title: str, setup_string: str, string: str, repeat: int = 1, number: int = 1,
                  width_title: int = 20, width_result: int = 10):
    results = timeit.repeat(
        string,
        setup_string,
        repeat=repeat,
        number=number,
    )
    result = min(results)
    print(f'{title:<{width_title}} {result/number:<{width_result},.05f} {number/result:<{width_result},.05f}')


def benchmark(*, skip_lib_comparisons: bool = False):
    width_title = 50
    width_result = 20
    bench = partial(benchmark_one, width_title=width_title, width_result=width_result)

    print('Test machine: {0} {2} {3} {4} {5}'.format(*platform.uname()))
    print()
    print(f'{"Test lib":<{width_title}} {"Min sec execution":<{width_result}} {"Calls/sec":<{width_result}}')

    bench(
        'xlsxio read big file with types',
        READ_XLSXIO_SETUP.format(filename=f"'{FILE_XLSX_BIG}'"),
        READ_XLSXIO_STRING.format(with_types=True, default_type='str'),
        repeat=3,
        number=1,
    )

    bench(
        'xlsxio read big file without types',
        READ_XLSXIO_SETUP.format(filename=f"'{FILE_XLSX_BIG}'"),
        READ_XLSXIO_STRING.format(with_types=False, default_type='str'),
        repeat=3,
        number=1,
    )

    bench(
        'xlsxio read big file without types all in bytes',
        READ_XLSXIO_SETUP.format(filename=f"'{FILE_XLSX_BIG}'"),
        READ_XLSXIO_STRING.format(with_types=False, default_type='bytes'),
        repeat=3,
        number=1,
    )

    bench(
        'xlsxio read big file with types from memory',
        READ_XLSXIO_SETUP.format(filename=FILE_XLSX_BIG_MEMORY),
        READ_XLSXIO_STRING.format(with_types=True, default_type='str'),
        repeat=3,
        number=1,
    )

    bench(
        'xlsxio read small file with types',
        READ_XLSXIO_SETUP.format(filename=f"'{FILE_XLSX_SMALL}'"),
        READ_XLSXIO_STRING.format(with_types=True, default_type='str'),
        repeat=100,
        number=10,
    )

    bench(
        'xlsxio read small file without types',
        READ_XLSXIO_SETUP.format(filename=f"'{FILE_XLSX_SMALL}'"),
        READ_XLSXIO_STRING.format(with_types=False, default_type='str'),
        repeat=100,
        number=10,
    )

    bench(
        'xlsxio read small file without types all in bytes',
        READ_XLSXIO_SETUP.format(filename=f"'{FILE_XLSX_SMALL}'"),
        READ_XLSXIO_STRING.format(with_types=False, default_type='bytes'),
        repeat=100,
        number=10,
    )

    bench(
        'xlsxio read small file with types from memory',
        READ_XLSXIO_SETUP.format(filename=FILE_XLSX_SMALL_MEMORY),
        READ_XLSXIO_STRING.format(with_types=True, default_type='str'),
        repeat=100,
        number=10,
    )

    if skip_lib_comparisons:
        return

    # openpyxl

    bench(
        'openpyxl read big file',
        READ_OPENPYXL_SETUP.format(filename=f"'{FILE_XLSX_BIG}'"),
        READ_OPENPYXL_STRING,
        repeat=3,
        number=1,
    )

    bench(
        'openpyxl read big file from memory',
        READ_OPENPYXL_SETUP.format(filename=FILE_XLSX_BIG_MEMORY),
        READ_OPENPYXL_STRING,
        repeat=3,
        number=1,
    )

    bench(
        'openpyxl read small file',
        READ_OPENPYXL_SETUP.format(filename=f"'{FILE_XLSX_SMALL}'"),
        READ_OPENPYXL_STRING,
        repeat=100,
        number=10,
    )

    bench(
        'openpyxl read small file from memory',
        READ_OPENPYXL_SETUP.format(filename=FILE_XLSX_SMALL_MEMORY),
        READ_OPENPYXL_STRING,
        repeat=100,
        number=10,
    )

    # xlrd

    bench(
        'xlrd read big file',
        READ_XLRD_SETUP.format(filename=f"'{FILE_XLSX_BIG}'"),
        READ_XLRD_STRING,
        repeat=3,
        number=1,
    )

    bench(
        'xlrd read big file from memory',
        READ_XLRD_SETUP.format(filename=FILE_XLSX_BIG_MEMORY),
        READ_XLRD_STRING,
        repeat=3,
        number=1,
    )

    bench(
        'xlrd read small file',
        READ_XLRD_SETUP.format(filename=f"'{FILE_XLSX_SMALL}'"),
        READ_XLRD_STRING,
        repeat=100,
        number=10,
    )

    bench(
        'xlrd read small file from memory',
        READ_XLRD_SETUP.format(filename=FILE_XLSX_SMALL_MEMORY),
        READ_XLRD_STRING,
        repeat=100,
        number=10,
    )

    # sxl

    bench(
        'sxl read big file',
        READ_SXL_SETUP.format(filename=f"'{FILE_XLSX_BIG}'"),
        READ_SXL_STRING,
        repeat=3,
        number=1,
    )

    bench(
        'sxl read big file from memory',
        READ_SXL_SETUP.format(filename=FILE_XLSX_BIG_MEMORY),
        READ_SXL_STRING,
        repeat=3,
        number=1,
    )

    bench(
        'sxl read small file',
        READ_SXL_SETUP.format(filename=f"'{FILE_XLSX_SMALL}'"),
        READ_SXL_STRING,
        repeat=100,
        number=10,
    )

    bench(
        'sxl read small file from memory',
        READ_SXL_SETUP.format(filename=FILE_XLSX_SMALL_MEMORY),
        READ_SXL_STRING,
        repeat=100,
        number=10,
    )

    # xlsx2csv

    bench(
        'xlsx2csv read big file',
        READ_OPENPYXL_SETUP.format(filename=f"'{FILE_XLSX_BIG}'"),
        READ_OPENPYXL_STRING,
        repeat=3,
        number=1,
    )

    bench(
        'xlsx2csv read big file from memory',
        READ_OPENPYXL_SETUP.format(filename=FILE_XLSX_BIG_MEMORY),
        READ_OPENPYXL_STRING,
        repeat=3,
        number=1,
    )

    bench(
        'xlsx2csv read small file',
        READ_OPENPYXL_SETUP.format(filename=f"'{FILE_XLSX_SMALL}'"),
        READ_OPENPYXL_STRING,
        repeat=100,
        number=10,
    )

    bench(
        'xlsx2csv read small file from memory',
        READ_OPENPYXL_SETUP.format(filename=FILE_XLSX_SMALL_MEMORY),
        READ_OPENPYXL_STRING,
        repeat=100,
        number=10,
    )


if __name__ == '__main__':
    benchmark(skip_lib_comparisons="skip-lib-comps" in sys.argv)
